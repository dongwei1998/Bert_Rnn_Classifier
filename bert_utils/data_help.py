# -*- coding:utf-8 -*-
# @Time      : 2021-08-02 15:57
# @Author    : 年少无为呀！
# @FileName  : data_help.py
# @Software  : PyCharm
import random
import csv
import openpyxl
from parameter import args

def excel2csv_data_create(args):
    train_writer = csv.writer(open(args.train_csv, 'w', encoding='utf-8', newline=''))
    dev_writer = csv.writer(open(args.dev_csv, 'w', encoding='utf-8', newline=''))
    test_writer = csv.writer(open(args.test_csv, 'w', encoding='utf-8', newline=''))
    excel_data = openpyxl.load_workbook(args.org_files_xlsx)
    table_names = excel_data.sheetnames
    table = excel_data[table_names[2]]
    max_row = table.max_row
    for idx in range(2, max_row):
        text = table.cell(idx, 1).value
        label = table.cell(idx, 3).value
        # class_dict = {'胜诉': 0, '其他': 1, '撤诉': 2, '败诉': 3}
        # label = class_dict[label]
        line = [label, text]
        _rand = random.random()
        if _rand < 0.80:
            # 训练数据
            train_writer.writerow(line)
        elif _rand < 0.99:
            # 验证数据
            dev_writer.writerow(line)
        else:
            # 测试数据
            test_writer.writerow(line)
